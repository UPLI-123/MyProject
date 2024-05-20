import tkinter as tk
from openpyxl import load_workbook, Workbook
from tkinter import ttk
from tkinter import filedialog
import tkinter.messagebox
import sqlite3
import os
import shutil # 用于文件的复制

# TODO 设置模板布局设计
class set_UI():
    # 初始化
    def __init__(self,root,loop,name,user_info):
        self.root = root
        self.loop =loop
        self.name =name
        self.user_info = user_info
        # 显示布局
        self.display()
        pass
    #  todo 退出操作的实现
    def method_exit(self):
        self.loop.destroy()
        pass
    # todo 导入用户信息
    def method_import_user(self):
        # 导入用户 操作
        path = tk.filedialog.askopenfilename(title = '请选择导出的位置',filetypes=[("数据库文件", ".db")],defaultextension='.db')
        if not path:
            return
        try:
            path_name = path.split('/')[-1]
            # print(path)
            # print(path_name)
            cur_path = os.getcwd()+"/"+path_name
            shutil.copyfile(path,cur_path)

            with sqlite3.connect("{}".format(path_name)) as conn , sqlite3.connect("User.db") as coon:
                cur01 = conn.cursor()
                cur = coon.cursor()
                sql01 = "select *from login"
                cur01.execute(sql01)
                username,password,count,image = cur01.fetchall()[0]
                sql = "insert into login values('{}','{}')".format(username,password)
                cur.execute(sql)
                pass
            tk.messagebox.showinfo("提示", "导入成功")
        except:
            tk.messagebox.showinfo("提示","用户已存在，请先注销 再导入")



        pass
    # todo 导出用户信息
    def method_destroy(self):
        # 大致思想是 将 数据库进行粘贴复制 ，然后将数据库进行备份。
        if self.name == 'admin':
            tk.messagebox.showinfo("提示","该数据库不支持导出")
            return
            pass
        # 选择 导出位置
        path = tk.filedialog.asksaveasfilename(title = '请选择导出的位置',filetypes=[("数据库文件", ".db")],initialfile = self.name,defaultextension='.db')
        # print(path)
        # 接下来实现 一个文件的复制到另一个地方上去
        # 获得当前的工作目录
        if not path:
            return
        cur_path = os.getcwd()
        # print(cur_path)
        # 需要备份的文件
        backup_file = cur_path+'\\'+self.name+".db"
        # 实现文件的复制操作
        shutil.copyfile(backup_file,path)
        # 最后显示 导出成功的通知
        tk.messagebox.showinfo("提示","导出数据库成功")
        pass
    # todo 创建用户信息
    def method_create(self):
        # 判断信息同登录时， 的信息判断
        username = self.var_user.get()
        # 对用户名进行判断
        if username == "" or (not username):
            tk.messagebox.showwarning("警告", "用户名不能为空")
            # 将光标定义在用户名的文本框中
            self.en_user.focus()
            return
            pass
        # 获取密码
        password = self.var_pad01.get()
        # 对密码的格式进行判断
        if password == "" or (not password):
            tk.messagebox.showwarning("警告", "密码不能为空")
            # 将光标定义在密码的文本框中
            self.en_pad01.focus()
            return
        password02 = self.var_pad02.get()
        # 对密码的格式进行判断
        if password02 == "" or (not password02):
            tk.messagebox.showwarning("警告", "第二次密码不能为空")
            # 将光标定义在密码的文本框中
            self.en_pad02.focus()
            return
        # 用户名格式的判断
        if len(username) > 10 or len(username) < 3:
            tk.messagebox.showwarning("提示", "请输入正确的用户名")
            # 清空用户框
            self.var_user.set("")
            # 固定光标
            self.en_user.focus()
            return
            pass
        # 密码的格式判断
        if len(password) < 6 or len(password) > 18:
            tk.messagebox.showwarning("提示", "密码不符合规范")
            # 清空密码栏
            self.var_pad01.set("")
            # 定义光标
            self.en_pad01.focus()
            return
            pass

        # 验证两次密码是否一致
        if password02!=password:
            tk.messagebox.showinfo("提示","密码不一致请重新检测你的密码")
            return
        # 检测 数据库中是否已经有了该用户了
        with sqlite3.connect("User.db") as coon:
            cur = coon.cursor()
            sql = "select *from login where username = '{}'".format(username)
            cur.execute(sql)
            user_name = cur.fetchone()
            if not user_name:
                sql = "insert into login values('{0}','{1}')".format(username,password)
                cur.execute(sql)
                pass
            else:
                tk.messagebox.showinfo("提示","该用户名已近存在了")
                self.var_user.set("")
                self.en_user.focus()
                return
        #     额外的创建一个数据库用来存储信息
        with sqlite3.connect("{}.db".format(username)) as coon:
            cur = coon.cursor()
            sql = "create table login(username TEXT primary key,password TEXT,count integer,image TEXT)"
            # 用来暂存数据
            sql1 = "insert into login values ('{0}','{1}',0,'Head/tx.gif')".format(username,password)
            # 创建一个单词本表
            sql2 = "create table words(wid integer PRIMARY KEY AUTOINCREMENT,word text,means text,tag integer)"
            # 创建 一个临时表用来记录用户目前学习单词的情况
            sql3 = "create table temp(bookname text , nowlength integer)"
            cur.execute(sql)
            cur.execute(sql1)
            cur.execute(sql2)
            cur.execute(sql3)
        # 最后 ，提示创建用户成功
        tk.messagebox.showinfo("提示","创建用户成功")
        # 创建成功后将一切输入框进行清空
        self.var_pad01.set("")
        self.var_user.set("")
        self.var_pad02.set("")

        # pass
    # todo 更改用户信息
    def method_upd(self):
        #  首先  获得  用户的名称
        username = self.var_user.get()
        pad01 = self.var_pad01.get()
        pad02 = self.var_pad02.get()

        if not username:
            tk.messagebox.showinfo("提示","用户名不能为空")
            self.en_user.foucs()
            return
            pass
        if not pad01:
            tk.messagebox.showinfo("提醒","原密码不能为空")
            self.en_pad01.foucs()
            return
            pass
        if not pad02:
            tk.messagebox.showinfo("提醒","新密码不能为空")
            self.en_pad02.focus()
            return
            pass

        # 判断用户名是否存在

        with sqlite3.connect("User.db") as coon:
            cur = coon.cursor()
            sql = "select  * from login where username = '{}'".format(username)
            cur.execute(sql)
            tmp = cur.fetchall()[0]
            pass
        # print(user_name,user_password)

        # 账户的判断
        if not tmp:
            tk.messagebox.showinfo("提醒","该账户的信息不存在")
            self.var_user.set("")
            self.en_user.focus()
            return
            pass
        # 原密码的判断
        if tmp[1] != pad01:
            tk.messagebox.showinfo("提醒","该用户的原密码不正确")
            self.var_pad01.set("")
            self.en_pad01.focus()
            return
            pass

        # 新密码的格式判断
        if len(pad02) < 6 or len(pad02) > 18:
             tk.messagebox.showwarning("提示", "密码不符合规范")
             # 清空密码栏
             self.var_pad02.set("")
             # 定义光标
             self.en_pad02.focus()
             return
             pass

        # 更改成功 的操作
        # 更改操作 ， 以及相应的数据库操作
        with sqlite3.connect("User.db") as conn:
            cur = conn.cursor()
            sql = "update login set password = '{}' where username = '{}'".format(pad02,username)
            cur.execute(sql)
            pass
        # 同时也要更改数据库中的信息
        with sqlite3.connect("{}.db".format(username)) as conn:
            cur = conn.cursor()
            sql = "update login set password = '{}' where username = '{}'".format(pad02,username)
            cur.execute(sql)
            pass
        tk.messagebox.showinfo("提示","修改成功")
        # 清空
        self.var_user.set("")
        self.var_pad01.set("")
        self.var_pad02.set("")
        self.en_user.focus()
        pass
    #  todo 重置功能的实现
    def method_reset(self):
        # 变量设置为空
        self.var_search.set("")
        # 记录信息为空
        self.var_book_id.set("")
        # 提示信息
        tk.messagebox.showinfo("提示", "重置成功")
        # 重新显示数据
        self.show_data()
        # 光标到达搜索框中
        self.en_search.focus()
        pass
    # todo 鼠标点击 事件
    def method_click(self):
        # 如果表单没有被选中，那么就什么也不显示
        if not self.treeAddressList.selection():
            return
        # 获得 每一行的编号
        item = self.treeAddressList.selection()[0]
        # 将搜索栏中显示数据
        self.var_search.set(self.treeAddressList.item(item,'values')[0])
        self.en_search.focus() # 聚焦
        # 查询 相应单词书的编号
        with sqlite3.connect('User.db') as coon:
            cur = coon.cursor()
            sql = "select bid from book where bname = '{}'".format(self.var_search.get())
            cur.execute(sql)
            bid = cur.fetchall()[0][0]
            pass
        self.var_book_id.set(bid)
        # print(self.var_book_id.get())
        # print(item)
        pass
    # todo 修改功能的实现
    def method_update(self):
        # 根据bid进行修改
        # 获取bid
        bid  = self.var_book_id.get()
        # 如果一开始没有选择的话
        if not bid:
            tk.messagebox.showwarning("提示","请选择你要修改的图书")
            return
        # 修改操作
        # 首先从文本框中获得修改后的内容
        new_name = self.var_search.get()
        # 修改book表中的名称，同时为了保持一致还要修改对应名称表的名称
        with sqlite3.connect("User.db") as coon:
            cur = coon.cursor()
            sql = "select bname from book where bid = {}".format(bid)
            cur.execute(sql)
            old_name = cur.fetchone()[0]
            if old_name==new_name:
                return
            # print(old_name)
            sql01 = "update book set bname = '{0}' where bid = {1}".format(new_name,bid)
            sql02 = "alter table {0} rename to {1}".format(old_name,new_name)
            cur.execute(sql01)
            cur.execute(sql02)
            pass
        # 提示
        tk.messagebox.showinfo("提示","修改成功")
        # 显示数据
        self.show_data()
        pass
    # todo 导出功能的实现
    def method_out(self):
        # 首先 ，先获得要导出的书籍的名称
        book_name = self.var_search.get()
        if not book_name:
            tk.messagebox.showinfo("提示","请选择你要导出的书籍")
            return
        # 首先从数据库中将书籍的信息取出来
        with sqlite3.connect("User.db") as conn:
            cur = conn.cursor()
            sql = "select * from {} ".format(book_name)
            cur.execute(sql)
            wordslist = cur.fetchall()
            pass
        # 创建保存文件的位置,指定默认的保存类型
        try:
            path = tk.filedialog.asksaveasfilename(title=u'保存文件',defaultextension='.xlsx', filetypes=[('Excel', '.xlsx')],initialfile =book_name)
            # print(path)
            print(wordslist)
            # 创建一个工作本
            workbook = Workbook()
            # 创建一个默认sheet
            sheet = workbook.active
            sheet['A1'] = '单词'
            sheet['B1'] = '翻译'
            length = len(wordslist)
            start  =2
            x = 0
            for row in range(start,length+2):
                sheet.cell(row,1,wordslist[x][0])
                sheet.cell(row,2,wordslist[x][1])
                x += 1
            workbook.save(path)
            tk.messagebox.showinfo("提示","导入成功")
        except Exception:
            return
            pass
        pass
    # todo 学习此书功能的实现
    def method_learn_book(self):
        if self.name == 'admin':
            tk.messagebox.showinfo("提示","请先选择用户")
            return
        book_name = self.var_search.get()
        if not book_name:
            return
        with sqlite3.connect("User.db") as coon :
            cur = coon.cursor()
            sql = "select *from book where bname = '{}'".format(book_name)
            cur.execute(sql)
            book_list = cur.fetchall()
            pass
        if not book_list:
            tk.messagebox.showinfo("提示","请选择正确的书籍名称")
            return
        # 通过数据库来实现该功能
        with sqlite3.connect("User.db") as coon :
            cur = coon.cursor()
            # 先将其他的都置位 0
            sql01 = "update book set version = 0"
            sql = "update book set version =1 where bname = '{}'".format(book_name)
            cur.execute(sql01)
            cur.execute(sql)
            pass
        # 同时也将 该书籍 写到 用户数据库中
        with sqlite3.connect("{}.db".format(self.name)) as coon:
            cur = coon.cursor()
            sql01 = "select bookname from temp"
            cur.execute(sql01)
            b_name = cur.fetchall()
            if tuple((book_name,)) not in b_name:
                sql = "insert into temp values('{}',0)".format(book_name)
                cur.execute(sql)
            pass
        tk.messagebox.showinfo("提示","设置目标成功，去单词学习吧")
        # 接下来 来更新 正在学习的书籍
        self.user_info.configure(text = book_name)

        pass

    #  todo 模糊查询 查询要学习的书籍
    def method_findBook(self):
        # 获取 文本框中的名字
        book_name = self.var_search.get()
        # 打开数据库进行模糊查询
        with sqlite3.connect("User.db") as conn:
            cursor = conn.cursor()
            sql = "select * from book where bname like "+"'%"+book_name+"%'"
            cursor.execute(sql)
            book_info = cursor.fetchall()
            # print(book_info)
            pass
        # 将数据添加到表单中
        self.insert_data(book_info)

        pass
    #  todo 导入书籍方法
    def method_imp(self):
        # 弹出选择框进行选择的文件
        path = filedialog.askopenfilename()
        if not path:
            return
        # 首先要对 传入的文件进行判断看是否为Excel文件
        # 先获得相应的文件名及其后缀名
        file_name = path.split("/")[-1]
        # print(file_name)
        # 判断字符串的后缀
        if file_name.split(".")[-1]!="xlsx":
            tk.messagebox.showerror("错误","所导入的文件必须是.xlsx文件")
            return
        # 进行数据库的导出操作
        # 文件名就是单词书的名称
        book_name = file_name.split(".")[0]
        # print(book_name)
        # 首先先从Excel中获取 数据
        # 加载 要读取的文档
        file = load_workbook(path)
        # 获得所有标签页的名称
        # print(file.sheetnames)
        # 获取存储数据的那一个sheet
        sheet_name = file.sheetnames[-1]
        # print(sheet)
        # 获取指定的sheet
        print(sheet_name)
        sheet  = file[sheet_name]
        # 使用len来获得该表格的全部的行数
        # print(sheet.rows)
        book_length = 0
        # 遍历每一行，然后将所遍历到的数据放到一个列表中
        book_list = list()
        for row in sheet.rows:
            book_length +=1
            tmp_list = list()
            for item in row:
                tmp_list.append(item.value)
            book_list.append(tmp_list)
        # print(book_length)
        # print(book_list)
        # 求得所包含的单词量
        book_length = book_length -1
        try:

            # 打开数据库,对数据库进行操作
            with sqlite3.connect("User.db") as conn:
                # 获得游标
                cursor = conn.cursor()
                # 首先将 单词书的名称和 所包含的单词的数量 填入相应的数据库
                sql = "insert into book(bname,bwords,version) values ("+"'"+book_name+"',"+str(book_length)+",0"\
                      +")"
                # 执行
                cursor.execute(sql)
                # 接下来进行create 一个单词书的表格
                sql01 = "create table "+book_name+"(word Text,means Text)"
                cursor.execute(sql01)
                # 将列表中数据放入数据库中
                for words in book_list[1::]:
                    cursor.execute("insert into "+ book_name + " values("+"'"+words[0]+"',"+"'"+words[1]+"')")
                    pass
                # 提交事务
                conn.commit()
                tk.messagebox.showinfo("通知","导入数据成功！！！")
        except:
            tk.messagebox.showinfo("通知","导入失败")
            # 显示表格
        self.show_data()
        pass

        pass

    # 插入数据到表单中的操作
    def insert_data(self,temp):
        # 首先 把表中原来的 空行进行清空
        for row in self.treeAddressList.get_children():
            self.treeAddressList.delete(row)
        # 将获取到的数据插入到表中
        for i, item in enumerate(temp):
            self.treeAddressList.insert("", i, values=item[1:])
        # 将字体变大
        # 首先获得相应的字列表
        items = self.treeAddressList.get_children()
        for item in items:
            self.treeAddressList.item(item, tags='oddrow')  # 对每一个单元格命名
            # 对子列表的进行字体和字号的设置
            self.treeAddressList.tag_configure('oddrow', font='宋体 20')
        pass

    # todo 显示数据界面
    def show_data(self):
        # 读取数据库中的信息
        with sqlite3.connect('User.db') as conn:
            cur = conn.cursor()
            cur.execute("select *from book")
            # 获取数据
            temp = cur.fetchall()
            # print(temp)
            pass
        self.insert_data(temp)
        pass


    # 定义布局
    def display(self):
        # 基本思想 ， 首先将 一个 用于用户信息的注册 ， 另个一个用于单词书的导入，所以需要俩个框架。
        # 第一个单词书的导入
        self.frame_imBook = tk.LabelFrame(self.root,text = '单词书',font=('宋体',24))
        self.frame_imBook.place(relx=0.0,rely=0.0, relwidth=1, relheight=0.7)
        # 首先支持模糊查询进行查找相应的单词书
        # 书名标签
        self.lb_book_name = tk.Label(self.frame_imBook,text = '书 名:',font = ('宋体',22))
        self.lb_book_name.place(relx=0.1,rely=0.1, relwidth=0.1, relheight=0.1)
        # 搜索框
        # 变量的绑定
        self.var_search = tk.StringVar()
        self.var_search.set("")
        # 创建一个变量用来存储单词书前面的编号
        self.var_book_id = tk.StringVar()
        self.var_book_id.set("")
        self.en_search = tk.Entry(self.frame_imBook,font = ('宋体',22),textvariable = self.var_search)
        self.en_search.place(relx=0.22,rely=0.1, relwidth=0.3, relheight=0.1)
        # 查询 和 修改 重置 按钮
        self.but_search = tk.Button(self.frame_imBook,text = '查 询',font = ('宋体',22),command = self.method_findBook)
        self.but_search.place(relx=0.56,rely=0.1, relwidth=0.1, relheight=0.1)
        self.but_update = tk.Button(self.frame_imBook,text = '修 改',font = ('宋体',22),command = self.method_update)
        self.but_update.place(relx=0.70, rely=0.1, relwidth=0.1, relheight=0.1)
        self.but_reset = tk.Button(self.frame_imBook,text = '重 置',font = ('宋体',22),command = self.method_reset)
        self.but_reset.place(relx=0.83, rely=0.1, relwidth=0.1, relheight=0.1)
        # 通过style方法创建表单的格式
        self.style = ttk.Style()
        self.style.configure("Treeview.Heading",font=('宋体',22))
        # 通过style01来设置行高
        self.style01 = ttk.Style()
        # 通过rowheight属性来设计行高为40
        self.style01.configure("Treeview",rowheight=40)
        # 表单的设计,使用 TreeAddressList组件 进行实现
        # 首先 创建 一个 frame 来存放表单
        self.frame_table = tk.Frame(self.frame_imBook)
        # 放置的位置
        self.frame_table.place(relx=0.1, rely=0.25, relwidth=0.83, relheight=0.55)
        # 将滚定条放置到frame上
        self.scorllBar = tk.Scrollbar(self.frame_table)
        self.scorllBar.pack(side = tk.RIGHT,fill= tk.Y)
        # 使用TreeAddressList组件进行存放数据
        # 将组件放到frame上 ，同时 设置列头 ，以及同滚动条绑定
        self.treeAddressList = ttk.Treeview(self.frame_table,columns =('c1','c2'),\
                                            show="headings",\
                                            yscrollcommand = self.scorllBar.set)
        # 设置 每一列的格式
        self.treeAddressList.column('c1',width = 400 ,anchor = 'center')
        self.treeAddressList.column('c2',width = 339, anchor = 'center')
        # 设置列头
        self.treeAddressList.heading('c1',text = '书名')
        self.treeAddressList.heading('c2',text = '单词数')
        # 将表单显示出来
        self.treeAddressList.pack(side = tk.LEFT ,fill = tk.Y)
        self.scorllBar.config(command = self.treeAddressList.yview)
        # 将表单显示出来
        self.show_data()
        # 将表单进行绑定
        self.treeAddressList.bind('<Button-1>',lambda event:self.method_click())
        # 导入书籍 导出书籍 学习书籍
        # 导入书籍
        self.but_imBook = tk.Button(self.frame_imBook,text = '导入书籍' ,font = ('宋体',22),command = self.method_imp)
        self.but_imBook.place(relx=0.1, rely=0.85, relwidth=0.20, relheight=0.1)
        # 学习书籍
        self.but_learn = tk.Button(self.frame_imBook,text = '学习书籍' ,font= ('宋体',22),command = self.method_learn_book)
        self.but_learn.place(relx=0.42, rely=0.85, relwidth=0.20, relheight=0.1)
        # 导出书籍
        self.but_outBook = tk.Button(self.frame_imBook,text = '导出书籍' ,font = ('宋体',22),command=self.method_out)
        self.but_outBook.place(relx=0.73, rely=0.85, relwidth=0.20, relheight=0.1)

        # 第二个模块用户信息模块
        self.frame_user = tk.LabelFrame(self.root,text = '用 户',font =("宋体",24))
        self.frame_user.place(relx=0.0,rely=0.7,relwidth =1 , relheight = 0.3)
        # 用户分为两个界面
        # 当前用户阶段
        self.frame_now_user = tk.Frame(self.frame_user)
        self.frame_now_user.place(relx=0.0,rely=0.0,relwidth =0.3 , relheight = 1)
        # 首先是一个标签
        self.lable_now_user = tk.Label(self.frame_now_user,text = '当前用户：',font = ('宋体',20))
        self.lable_now_user.place(relx=0.0,rely=0.0,relwidth =0.5 , relheight = 0.38)
        # 用一个标签来显示用户名
        self.lable_user_name = tk.Label(self.frame_now_user,text = self.name,font = ('微软雅黑',20))
        self.lable_user_name.place(relx=0.5,rely=0.0,relwidth =0.5 , relheight = 0.38)
        # 创建一个frame 来进行导入用户和注销用户操作
        self.frame_op = tk.Frame(self.frame_now_user)
        self.frame_op.place(relx=0.0,rely=0.38,relwidth =1 , relheight = 0.62)
        # 接下来创建连个按钮操作
        self.but_import_user = tk.Button(self.frame_op,text = '导入用户信息' , font = ('宋体',18),command = self.method_import_user)
        self.but_import_user.place(relx=0.0,rely=0.0,relwidth =1 , relheight = 0.5)
        self.but_drop_user = tk.Button(self.frame_op, text='导出用户信息', font=('宋体', 18),command = self.method_destroy)
        self.but_drop_user.place(relx=0.0, rely=0.5, relwidth=1, relheight=0.5)

        # 注册用户和撤销用户登录界面
        self.frame_register = tk.Frame(self.frame_user)
        self.frame_register.place(relx=0.3,rely=0.0,relwidth =0.7 , relheight = 1)
        # 分为 两部分 ， 一部分是输入框，另一部分是登录 、退出 、 更改 操作
        self.frame_left = tk.Frame(self.frame_register)
        self.frame_left.place(relx=0.0,rely=0.0,relwidth =0.6 , relheight = 1)
        # 用户名 密码 密码
        self.lab_user  = tk.Label(self.frame_left,text = '用户名：',font =('宋体',20))
        self.lab_user.place(relx=0.05,rely=0.15,relwidth =0.3 , relheight = 0.2)
        self.var_user = tk.StringVar()
        self.var_user.set("")
        self.en_user = tk.Entry(self.frame_left,font =('宋体',20),textvariable = self.var_user)
        self.en_user.place(relx=0.4,rely=0.15,relwidth =0.4 , relheight = 0.2)
        #
        self.lab_pad01 = tk.Label(self.frame_left, text='密 码 ：', font=('宋体', 20))
        self.lab_pad01.place(relx=0.05, rely=0.43, relwidth=0.3, relheight=0.2)
        self.var_pad01 = tk.StringVar()
        self.var_pad01.set("")
        self.en_pad01 = tk.Entry(self.frame_left, font=('宋体', 20), textvariable=self.var_pad01,show = '*')
        self.en_pad01.place(relx=0.4, rely=0.43, relwidth=0.4, relheight=0.2)
        #
        self.lab_pad02 = tk.Label(self.frame_left, text='密 码 ：', font=('宋体', 20))
        self.lab_pad02.place(relx=0.05, rely=0.75, relwidth=0.3, relheight=0.2)
        self.var_pad02 = tk.StringVar()
        self.var_pad02.set("")
        self.en_pad02 = tk.Entry(self.frame_left, font=('宋体', 20), textvariable=self.var_pad02,show ='*')
        self.en_pad02.place(relx=0.4, rely=0.75, relwidth=0.4, relheight=0.2)

        self.frame_right = tk.Frame(self.frame_register)
        self.frame_right.place(relx=0.6,rely=0.0,relwidth =0.4, relheight = 1)
        # 排放按钮
        self.but_login = tk.Button(self.frame_right,text="创 建",font= ('宋体',20),command = self.method_create)
        self.but_login.place(relx=0.0, rely=0.0, relwidth=1, relheight=0.33)

        self.but_upd = tk.Button(self.frame_right, text="更 改", font=('宋体', 20),command = self.method_upd)
        self.but_upd.place(relx=0.0, rely=0.33, relwidth=1, relheight=0.33)

        self.but_exit = tk.Button(self.frame_right, text="退 出", font=('宋体', 20),command = self.method_exit)
        self.but_exit.place(relx=0.0, rely=0.66, relwidth=1, relheight=0.34)
        pass
    pass