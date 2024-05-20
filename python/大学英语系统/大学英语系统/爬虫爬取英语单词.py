# 从扇贝英语上爬取相应的单词书,使用xpath来进行获取
import requests
from lxml import etree
from openpyxl import Workbook,load_workbook
# 使用 python 爬取 第一个 python单词 书
# 扇贝单词 python 词汇的接口网站
# https://www.shanbay.com/wordlist/104899/202159/?page=1
# 准备 请求头
headers = {
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.66 Safari/537.36',
'cookie': 'csrftoken=LtXOGSNQusRryMaL51LcqwdksBHdRk5p; _ga=GA1.2.260363577.1600949302; __utmc=183787513; abTestVersion=A; __utmz=183787513.1607483972.5.5.utmcsr=web.shanbay.com|utmccn=(referral)|utmcmd=referral|utmcct=/; sessionid="e30:1kmq1y:983QyvwPZomxfITYZiln1bVz0tI"; auth_token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpZCI6MjI4OTE3NzYxLCJleHAiOjE2MDgzNDg0MjQsImV4cF92MiI6MTYwODM0ODQyNCwiZGV2aWNlIjoiIiwidXNlcm5hbWUiOiJRcV9hODRmMjAzMDgyNDRmZGI2IiwiaXNfc3RhZmYiOjAsInNlc3Npb25faWQiOiJiNWVlOThkODM5Y2QxMWViOTdkOTNlMzE5MzQ5YzVlYSJ9.Jm1h8uBGDzi9dWLFpHCopFPt4YIBAefMuwlaBu4ZQGw; userid=228917761; __utma=183787513.260363577.1600949302.1607483972.1607510984.6; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%22dbylmv%22%2C%22first_id%22%3A%22174c003be3d102-07bdb4fc1c17ac-333769-1327104-174c003be3e265%22%2C%22props%22%3A%7B%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%22%2C%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC%22%2C%22%24latest_utm_source%22%3A%22web_codetime%22%2C%22%24latest_utm_medium%22%3A%22shanbay_nav%22%2C%22%24latest_referrer_host%22%3A%22www.baidu.com%22%7D%2C%22%24device_id%22%3A%22174c003be3d102-07bdb4fc1c17ac-333769-1327104-174c003be3e265%22%7D; __utmt=1; __utmb=183787513.17.10.1607510984'
}
url = "https://www.shanbay.com/wordlist/104899/202162/?page="

# 首先 先打开 要导入数据 的 Excel 文件
wb = Workbook() # 文件名
# #  创建一个分页
ws = wb.create_sheet()
# 写好 每一个  名称
ws.cell(1,1,'单词')
ws.cell(1,2,'翻译')
# 定义开始的时候的单词的长度
length = 0
# 定义开始的下标
start = 2

# 爬取9 页的内容
for i in range(1,10):
    url1 = url+str(i)
    print(url1)
    response = requests.get(url=url1, headers=headers)
    # 查看请求是否能够获得成功
    # print(response) # <Response [200]> 获取成功
    # 看看是否有中文乱码
    # print(response.text) # 没有中文乱码
    # 使用xpath 进行爬取数据
    # 先将字符串转化为html文件
    root = etree.HTML(response.text)
    # 分析路径进行爬取响应的单词
    words = root.xpath('//td[@class="span2"]/strong/text()')
    #  分析路径爬取相应的解释
    means = root.xpath('//td[@class="span10"]/text()')
    # 先写到一个excel文件中来
    # 用来记录下标
    x = 0
    length = length + len(words)
    for row in range(start,length+2):
            ws.cell(row,1,words[x])
            ws.cell(row,2,means[x])
            x = x+1
            pass
    start = length + 2
    pass
## 保存文件
wb.save("python")












