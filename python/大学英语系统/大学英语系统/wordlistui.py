import tkinter as tk
from tkinter import ttk
import sqlite3
from tkinter import messagebox
import pyttsx3
from openpyxl import load_workbook,Workbook
import time

# 单词本界面设计
class wordlist_ui():
    def __init__(self,root,name):
        self.root = root
        self.name = name
        self.display()
        pass
    # todo list操作
    def m_list(self):
        # 单词本框架
        self.frame_op = tk.Frame(self.root)
        self.frame_op.place(relx=0, rely=0.1, relheight=0.9, relwidth=1)
        self.wordlist_frame = tk.Frame(self.frame_op)
        self.wordlist_frame.place(relx=0, rely=0.0, relheight=0.9, relwidth=0.4)
        # 将单词显示出来
        self.table_tree()
        self.left_frame = tk.Frame(self.frame_op)
        self.left_frame.place(relx=0.4, rely=0.0, relheight=0.9, relwidth=0.6)

        self.t_frame = tk.Frame(self.left_frame, bg="#fafafa")
        self.t_frame.place(relx=0.0, rely=0.0, relheight=0.4, relwidth=1)

        self.means_text = tk.Text(self.left_frame, font=('宋体', 22), bd=10)
        self.means_text.place(relx=0.0, rely=0.4, relheight=0.6, relwidth=1)
        pass
    # todo 移除操作
    def remove_word(self):
        try:
            if not self.list_word:
                return
        except:
            return
        # 首先从数据库 中移除
        with sqlite3.connect("{}.db".format(self.name)) as coon:
            cur = coon.cursor()
            sql = "delete from words where wid = {}".format(self.list_word[0][0])
            cur.execute(sql)
            pass
        messagebox.showinfo("提示","删除成功")
        # 将文本框也进行清空
        self.means_text.delete(0.0,tk.END)
        self.clean_label(self.t_frame)
        # 清空选中项
        self.list_word = []
        # 重新显示表单
        self.show_data(0)
        pass
    # todo 升序操作
    def method_asc(self):
        self.show_data(1)
        pass
    # todo 降序操作
    def method_desc(self):
        self.show_data(2)
        pass
    # todo 导出操作
    def method_imp(self):
        # 首先从数据库中将书籍的信息取出来
        with sqlite3.connect("{}.db".format(self.name)) as conn:
            cur = conn.cursor()
            sql = "select * from words "
            cur.execute(sql)
            wordslist = cur.fetchall()
            pass
        # 创建保存文件的位置,指定默认的保存类型
        path = tk.filedialog.asksaveasfilename(title=u'保存文件', defaultextension='.xlsx', filetypes=[('Excel', '.xlsx')],
                                               initialfile="MyWords")
        # print(path)
        print(wordslist)
        # 创建一个工作本
        workbook = Workbook()
        # 创建一个默认sheet
        sheet = workbook.active
        sheet['A1'] = '单词'
        sheet['B1'] = '翻译'
        length = len(wordslist)
        start = 2
        x = 0
        for row in range(start, length + 2):
            sheet.cell(row, 1, wordslist[x][1])
            sheet.cell(row, 2, wordslist[x][2])
            x += 1
        workbook.save(path)
        tk.messagebox.showinfo("提示", "导出成功")
        pass
    # todo 恢复操作
    def method_recover(self):
        self.en_search.destroy()
        self.but_concel.destroy()
        self.but_search = tk.Button(self.on_frame, image=self.img_search, bg="#fff", bd=0, command=self.method_rearsh)
        self.but_search.place(relx=0.3, rely=0.1, relheight=0.4, relwidth=0.02)
        pass
    # todo 搜索实现操作
    def method_rearshImpl(self):
        # print(111111111111)
        if not self.var_search.get():
            self.show_data(0)
            pass
        else:
            self.show_data(3)
        pass
    #  todo 搜索操作
    def method_rearsh(self):
        # 首先点击 搜索 框 ，搜索框先消失，然后出现一个文本框
        self.but_search.destroy()
        # 然后创建一个 文本框
        self.var_search = tk.StringVar(self.on_frame)
        self.var_search.set("")
        self.en_search = tk.Entry(self.on_frame,relief =tk.GROOVE,textvariable = self.var_search)
        self.en_search.place(relx=0.3, rely=0.1, relheight=0.4, relwidth=0.2)
        # self.but_search.place(relx=0.3, rely=0.1, relheight=0.4, relwidth=0.02)
        self.img_concel = tk.PhotoImage(file ="image/6_02_03.gif")
        self.but_concel = tk.Button(self.on_frame,image = self.img_concel,bg ="#fff",bd = 0,command = self.method_recover)
        self.but_concel.place(relx=0.5, rely=0.1, relheight=0.4, relwidth=0.05)
        self.en_search.bind("<Return>",lambda event:self.method_rearshImpl())
        pass
    # todo 列表操作
    def method_list(self):
        # self.clean_label(self.wordlist_frame)
        # self.clean_label(self.left_frame)
        # 先清空，然后再重布置
        self.clean_label(self.frame_op)
        self.m_list()
        # 首先 设置一个排序窗口
        self.but_order_asc = tk.Button(self.on_frame, text='升序', fg="#666666", command=self.method_asc)
        self.but_order_desc = tk.Button(self.on_frame, text='降序', fg="#666666", command=self.method_desc)
        # 将两个 按钮放置好
        self.but_order_asc.place(relx=0.05, rely=0.1, relheight=0.4, relwidth=0.1)
        self.but_order_desc.place(relx=0.15, rely=0.1, relheight=0.4, relwidth=0.1)
        self.img_imp = tk.PhotoImage(file="image/6_01.gif")
        self.but_imp = tk.Button(self.on_frame, image=self.img_imp, bg="#fff", bd=0, anchor='w',command=self.method_imp)
        self.but_imp.place(relx=0.25, rely=0.1, relheight=0.4, relwidth=0.05)
        # 查找框的显示
        self.img_search = tk.PhotoImage(file="image/5_02.gif")
        self.but_search = tk.Button(self.on_frame, image=self.img_search, bg="#fff", bd=0, command=self.method_rearsh)
        self.but_search.place(relx=0.3, rely=0.1, relheight=0.4, relwidth=0.02)
        pass
    # todo 复习布局
    def m_rewirte(self):
        # 用个标签来存储当前所选单词书的名称
        self.speed_meet()
        pass
    # todo 复习操作
    def method_rewirte(self):
        self.clean_label(self.frame_op)
        self.but_order_asc.destroy()
        self.but_order_desc.destroy()
        self.but_imp.destroy()
        self.but_search.destroy()
        self.m_rewirte()
        pass
    # TODO 拼写验证
    def vary_spell(self):
        if self.list_word[0][1] == self.var_input.get():
            # 清空拼写框
            self.en_input.destroy()
            # 同时清空 文本框
            self.means_text.delete(0.0,tk.END)
            # 同时清空选中项
            self.list_word = []
            self.table_tree()
        pass
    # todo 退出拼写模式
    def exit_spell(self):
        # messagebox.showinfo("提示","已经成功退出")
        self.en_input.destroy()
        self.means_text.insert(tk.END,"\n"+self.list_word[0][1])
        self.table_tree()
        pass
    # todo 拼写操作
    def spell_word(self):
        try:
            if not self.list_word:
                return
        except:
            return
        # 首先先清空单词列表
        self.clean_label(self.wordlist_frame)
        # 将一个文本框放在正中间 ， 用于接收输入
        self.var_input = tk.StringVar()
        self.var_input.set("")
        self.en_input = tk.Entry(self.wordlist_frame,font = ('微软雅黑',24),textvariable = self.var_input)
        self.en_input.pack()
        self.en_input.focus()
        # 进行 方法的绑定
        self.en_input.bind("<Return>",lambda event:self.vary_spell())
        self.en_input.bind("<Escape>",lambda event:self.exit_spell())

        pass
    # 速记方法的绑定
    def method_speed(self):
        self.lab_word.configure(text=self.wordlist[self.i][1])
        self.lab_mean.configure(text=self.wordlist[self.i][2])
        print("111111111111")
        self.i+=1
        if self.i == len(self.wordlist):
            self.i =0
            pass
        pass

    # todo 速记操作
    def speed_meet(self):
        # 获取当前单词本的单词的数量
        length = len(self.wordlist)
        # print(length)
        i = 1 # 用来 记录第一个单词
        # 两个标签 一个是单词 一个是翻译
        self.lab_word = tk.Label(self.frame_op,text = self.wordlist[0][1],font=("微软雅黑",30))
        self.lab_word.place(relx=0, rely=0, relheight=0.5, relwidth=1)
        self.lab_mean = tk.Label(self.frame_op,text = self.wordlist[0][2],font =('微软雅黑',30))
        self.lab_mean.place(relx=0, rely=0.5, relheight=0.5, relwidth=1)
        self.i = 1
        self.frame_op.bind("<Return>",lambda event:self.method_speed())
        # 当前 框架被选中
        self.frame_op.focus_set()
        pass
    # todo 鼠标绑定事件
    def click_event(self):
        if not self.treeAddresslist.selection():
            return
        # 获得每一行的编号
        item = self.treeAddresslist.selection()[0]
        # 根据编号找到单词的名字
        word = self.treeAddresslist.item(item,"values")[0]
        # print(word)
        # 根据单词找到其他的信息
        with sqlite3.connect("{}.db".format(self.name)) as coon:
            cur = coon.cursor()
            sql ="select *from words where word = '{}'".format(word)
            cur.execute(sql)
            self.list_word = cur.fetchall()
            pass
        # print(self.list_word)
        # 先清空在显示
        self.clean_label(self.t_frame)
        # 首先 需要 一个用来接收单词
        self.word = tk.Label(self.t_frame,bg ="#fafafa",text = self.list_word[0][1],font =('微软雅黑',20,"bold"))
        self.word.place(relx=0.05, rely=0.2, relheight=0.2, relwidth=0.3)
        self.img_voice = tk.PhotoImage(file = "image/4_01_03.gif")
        self.but_voice  = tk.Button(self.t_frame,image = self.img_voice,command =self.sounds,bg ="#fafafa",bd = 0,anchor = 'w')
        self.but_voice.place(relx=0.45, rely=0.2, relheight=0.2, relwidth=0.1)
        # 显示出来删除操作
        self.img_del = tk.PhotoImage(file = "image/6_03.gif")
        self.but_remove = tk.Button(self.t_frame,image = self.img_del,command = self.remove_word,anchor = 'w',bd = 0 ,bg ="#fafafa")
        self.but_remove.place(relx=0.55, rely=0.2, relheight=0.2, relwidth=0.1)

        self.means_text.delete(0.0,tk.END)
        self.means_text.insert(tk.END,self.list_word[0][2])
        pass
    # todo 更改表中数据的大小
    def insert_data(self):
        # 首先  把表中原来的数据进行清空
        for row in self.treeAddresslist.get_children():
            self.treeAddresslist.delete(row)
        # 获取新的数据
        for i ,item in enumerate(self.wordlist):
            self.treeAddresslist.insert("",i,values = item[1])
            pass
        # 将字体进行变大
        items = self.treeAddresslist.get_children()
        for item in items:
            self.treeAddresslist.item(item,tags = 'row')
            self.treeAddresslist.tag_configure("row",font = '微软雅黑 20')
            pass
        pass
    # todo 显示单词本中的数据
    def show_data(self,count):
        if count == 0:
            with sqlite3.connect("{}.db".format(self.name)) as coon:
                cur = coon.cursor()
                sql = "select *from words"
                cur.execute(sql)
                self.wordlist = cur.fetchall()
                pass
        elif count == 1:
            with sqlite3.connect("{}.db".format(self.name)) as coon:
                cur = coon.cursor()
                sql = "select *from words order by word ASC"
                cur.execute(sql)
                self.wordlist = cur.fetchall()
                pass
        elif count == 2:
            with sqlite3.connect("{}.db".format(self.name)) as coon:
                cur = coon.cursor()
                sql = "select *from words order by word DESC"
                cur.execute(sql)
                self.wordlist = cur.fetchall()
                pass
        elif count == 3:
            with sqlite3.connect("{}.db".format(self.name)) as coon:
                cur = coon.cursor()
                sql = "select *from words where word like '%{}%'".format(self.var_search.get())
                cur.execute(sql)
                self.wordlist = cur.fetchall()
            pass

            pass
        # 将 表单中 的每一项的字符进行 修改,同时插入到treeview中
        self.insert_data()
        pass
    # todo 发音
    def sounds(self):
        try:
            if not self.list_word:
                return
        except:
            return
        gen = pyttsx3.init()
        gen.say(self.list_word[0][1])
        gen.say(self.list_word[0][2])
        gen.runAndWait()
        pass

    def table_tree(self):
        # 单词列表
        self.style_heading = ttk.Style()
        self.style_heading.configure("Treeview.Heading", font=('微软雅黑', 22))
        # 设置行高
        self.style01 = ttk.Style()
        self.style01.configure("Treeview", rowheight=40)
        self.scorllBar = tk.Scrollbar(self.wordlist_frame)
        self.scorllBar.pack(side=tk.RIGHT, fill=tk.Y)

        self.treeAddresslist = ttk.Treeview(self.wordlist_frame, columns=('c1'), \
                                            show="headings", \
                                            yscrollcommand=self.scorllBar.set)
        self.treeAddresslist.column('c1', width=400, anchor='center')

        self.treeAddresslist.heading('c1', text='单词')

        self.treeAddresslist.pack(side=tk.LEFT, fill=tk.Y)
        self.scorllBar.configure(command=self.treeAddresslist.yview)

        # 显示 单词本
        self.show_data(0)
        # 鼠标绑定事件
        self.treeAddresslist.bind('<Button-1>', lambda event: self.click_event())

    def display(self):

        # 将功能重新布局
        self.on_frame = tk.Frame(self.root,bg = "#fff")
        self.on_frame.place(relx=0, rely=0.0, relheight=0.1, relwidth=1)

        # 首先 设置一个排序窗口
        self.but_order_asc = tk.Button(self.on_frame,text = '升序',fg = "#666666",command = self.method_asc)
        self.but_order_desc = tk.Button(self.on_frame,text = '降序',fg = "#666666",command = self.method_desc)


        # 将两个 按钮放置好
        self.but_order_asc.place(relx=0.05, rely=0.1, relheight=0.4, relwidth=0.1)
        self.but_order_desc.place(relx=0.15, rely=0.1, relheight=0.4, relwidth=0.1)

        self.img_imp = tk.PhotoImage(file = "image/6_01.gif")

        self.but_imp = tk.Button(self.on_frame,image = self.img_imp,bg ="#fff",bd = 0,anchor = 'w',command = self.method_imp)
        self.but_imp.place(relx=0.25, rely=0.1, relheight=0.4, relwidth=0.05)

        # 查找框的显示
        self.img_search = tk.PhotoImage(file = "image/5_02.gif")

        self.but_search = tk.Button(self.on_frame,image = self.img_search,bg = "#fff",bd = 0,command = self.method_rearsh)
        self.but_search.place(relx=0.3, rely=0.1, relheight=0.4, relwidth=0.02)

        # 然后 是列表 、 卡片 、 复习
        self.img_list  = tk.PhotoImage(file = "image/5_03.gif")
        self.img_rewrite = tk.PhotoImage(file = "image/5_05.gif")

        # 三个按钮
        self.img_but_list = tk.Button(self.on_frame,image = self.img_list,bg="#fff",bd = 0,command = self.method_list)
        self.but_list = tk.Button(self.on_frame,text = '列表',fg = "#666",bd = 0,bg="#fff",anchor ='w',command = self.method_list)
        # self.img_but_cord = tk.Button(self.on_frame,image = self.img_cord,bg="#fff",bd = 0,command = self.method_cord)
        # self.but_cord = tk.Button(self.on_frame,text = '卡片',fg = "#666",bd = 0,bg="#fff",anchor ='w',command = self.method_cord)
        self.img_but_rewrite = tk.Button(self.on_frame,image = self.img_rewrite,bg="#fff",bd = 0,command = self.method_rewirte)
        self.but_rewrite = tk.Button(self.on_frame,text = '复习',fg = "#666",bd = 0,bg="#fff",anchor ='w',command = self.method_rewirte)

        self.img_but_list.place(relx=0.68, rely=0.1, relheight=0.4, relwidth=0.02)
        self.but_list.place(relx=0.7, rely=0.1, relheight=0.4, relwidth=0.05)

        # self.img_but_cord.place(relx=0.68, rely=0.1, relheight=0.4, relwidth=0.02)
        # self.but_cord.place(relx=0.7, rely=0.1, relheight=0.4, relwidth=0.05)

        self.img_but_rewrite.place(relx=0.76, rely=0.1, relheight=0.4, relwidth=0.02)
        self.but_rewrite.place(relx=0.78, rely=0.1, relheight=0.4, relwidth=0.04)
        # 复习后面 紧跟 一个 用来 记录 未学习单词数目
        # self.lab_word = tk.Label(self.on_frame,text = '0',fg = '#fff',anchor = 'w',bg = '#d20a0a')
        # self.lab_word.place(relx=0.82, rely=0.1, relheight=0.4, relwidth=0.05)
        self.m_list()


        # # 首先 一个大的标签先放到 单词本的最上
        # # 下方分为 两个  框架  ， 一个是单词本列表
        # # 四个按钮
        # self.frame_op = tk.Frame(self.other_frame,bg = 'red')
        # self.frame_op.place(relx=0, rely=0.4, relheight=0.6, relwidth=1)
        # #

        # #
        # self.but_spill = tk.Button(self.frame_op, text='拼写', font=('微软雅黑', 22),command = self.spell_word)
        # self.but_spill.place(relx=0.5, rely=0.5, relheight=0.5, relwidth=0.5)
        # #
        # self.but_scan = tk.Button(self.frame_op, text="速记", font=('微软雅黑', 22),command = self.speed_meet)
        # self.but_scan.place(relx=0.5, rely=0, relheight=0.5, relwidth=0.5)
        pass
    # 清空功能
    def clean_label(self,frame):
        for label in frame.winfo_children():
            label.destroy()
        pass
    pass
